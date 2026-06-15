"""
services/importacion.py
=======================
Plantilla Excel con dropdowns dependientes (Ruta filtrada por Cliente)
usando rangos nombrados + INDIRECT. Import hace match por nombre → ID.
"""

from __future__ import annotations

import io
import re
from datetime import date, time, datetime
from typing import BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from sqlalchemy.orm import Session

from ..models.ruta import Ruta
from ..models.vehiculo import Vehiculo
from ..models.cliente import Cliente
from ..models.paradero import Paradero
from ..schemas.servicio import ServicioCreate
from ..services.conductores import get_conductores_by_sede

# ── Estilos ───────────────────────────────────────────────────────────────────
COLOR_HEADER  = "1F3864"
COLOR_WHITE   = "FFFFFF"
COLOR_REQ     = "D9E1F2"
COLOR_OPT     = "EBF1DE"
COLOR_INFO    = "FFF2CC"
COLOR_REF_HDR = "2F5496"

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNAS = [
    # ── IDA ──────────────────────────────────────────────────────
    ("Fecha (DD/MM/AAAA)",           True),    # A
    ("Cliente",                      False),   # B
    ("Ruta",                         True),    # C
    ("Placa Vehículo IDA",           True),    # D
    ("Conductor 1 IDA",              True),    # E
    ("Conductor 2 IDA",              False),   # F
    ("Hora Salida IDA (HH:MM)",      True),    # G
    ("Hora Fin Manual (HH:MM)",      False),   # H
    ("Paradero de Origen",           False),   # I
    ("Observaciones",                False),   # J
    # ── RETORNO ──────────────────────────────────────────────────
    ("¿Misma unidad retorno?",       False),   # K  Sí / No
    ("Placa Vehículo RETORNO",       False),   # L  solo si K=No
    ("Conductor RETORNO",            False),   # M  solo si K=No
    ("Hora Salida Planta (HH:MM)",   False),   # N  cuando sale de planta
]

NUM_FILAS = 100   # Filas de datos en la plantilla


def _safe_name(texto: str) -> str:
    """Convierte un nombre en un identificador válido para rangos nombrados de Excel."""
    nombre = re.sub(r"[^A-Za-z0-9]", "_", texto.strip())
    if nombre and nombre[0].isdigit():
        nombre = "C_" + nombre
    return nombre[:31]   # Excel limita nombres a 255 chars, 31 es práctico


# ─────────────────────────────────────────────────────────────────────────────
# GENERAR PLANTILLA
# ─────────────────────────────────────────────────────────────────────────────
def generar_plantilla(sede_id: int, db: Session) -> bytes:
    wb = Workbook()

    # ── Cargar datos ──────────────────────────────────────────────────────────
    rutas = (
        db.query(Ruta)
        .filter(Ruta.sede_id == sede_id, Ruta.activa == True)
        .order_by(Ruta.nombre).all()
    )
    vehiculos = (
        db.query(Vehiculo)
        .filter(Vehiculo.sede_id == sede_id, Vehiculo.operativo == True)
        .order_by(Vehiculo.placa).all()
    )
    conductores = get_conductores_by_sede(sede_id, db)

    cliente_ids = list({r.cliente_id for r in rutas if r.cliente_id})
    clientes = (
        db.query(Cliente)
        .filter(Cliente.cliente_id.in_(cliente_ids), Cliente.activo == True)
        .order_by(Cliente.nombre).all()
    ) if cliente_ids else []

    paraderos = (
        db.query(Paradero)
        .filter(Paradero.sede_id == sede_id, Paradero.activo == True)
        .order_by(Paradero.nombre).all()
    )

    # ── Hojas de referencia (creadas ANTES de la hoja Importar) ──────────────
    ws_cli  = wb.active;  ws_cli.title = "Clientes"
    ws_rutas_hoja = wb.create_sheet("Rutas_Ref")
    ws_veh  = wb.create_sheet("Vehiculos")
    ws_cond = wb.create_sheet("Conductores")
    ws_par  = wb.create_sheet("Paraderos")

    _ref_clientes(ws_cli,  clientes)
    _ref_vehiculos(ws_veh, vehiculos)
    _ref_conductores(ws_cond, conductores)
    _ref_paraderos(ws_par, paraderos)

    # ── Hoja Rutas_Ref: una columna por cliente ───────────────────────────────
    # Estructura: fila 1 = encabezado (nombre cliente), filas 2+ = rutas
    # Esto permite crear rangos nombrados por cliente.
    rutas_por_cliente: dict[int, list[Ruta]] = {}
    for r in rutas:
        rutas_por_cliente.setdefault(r.cliente_id, []).append(r)

    # También una columna 1 con TODAS las rutas (fallback)
    _hdr_row(ws_rutas_hoja, ["TODAS LAS RUTAS"] + [c.nombre for c in clientes])

    # Llenar las columnas
    max_rutas_por_col = max((len(v) for v in rutas_por_cliente.values()), default=0)
    # Columna 1: todas las rutas
    for i, r in enumerate(rutas, start=2):
        ws_rutas_hoja.cell(row=i, column=1, value=r.nombre)
    # Columnas por cliente (columna 2 en adelante)
    for col_idx, cliente in enumerate(clientes, start=2):
        rutas_cliente = rutas_por_cliente.get(cliente.cliente_id, [])
        for row_idx, r in enumerate(rutas_cliente, start=2):
            ws_rutas_hoja.cell(row=row_idx, column=col_idx, value=r.nombre)

    # Formato de la hoja de referencia
    for ci in range(1, len(clientes) + 2):
        ws_rutas_hoja.column_dimensions[get_column_letter(ci)].width = 40
        c = ws_rutas_hoja.cell(row=1, column=ci)
        c.font = Font(bold=True, color=COLOR_WHITE)
        c.fill = PatternFill("solid", start_color=COLOR_REF_HDR)

    # ── Crear rangos nombrados para INDIRECT ──────────────────────────────────
    # Rango de todas las rutas
    n_rutas_total = max(len(rutas), 1)
    wb.defined_names["TODAS_RUTAS"] = DefinedName(
        "TODAS_RUTAS",
        attr_text=f"Rutas_Ref!$A$2:$A${n_rutas_total + 1}"
    )
    # Rango de clientes (para el dropdown de cliente)
    n_cli = max(len(clientes), 1)
    wb.defined_names["LISTA_CLIENTES"] = DefinedName(
        "LISTA_CLIENTES",
        attr_text=f"Clientes!$A$2:$A${n_cli + 1}"
    )
    # Rangos por cliente: nombre seguro como clave
    cliente_range_map: dict[int, str] = {}
    for col_idx, cliente in enumerate(clientes, start=2):
        rutas_cliente = rutas_por_cliente.get(cliente.cliente_id, [])
        if not rutas_cliente:
            continue
        range_name = _safe_name(cliente.nombre)
        n = len(rutas_cliente)
        attr = f"Rutas_Ref!${get_column_letter(col_idx)}$2:${get_column_letter(col_idx)}${n + 1}"
        # Evitar duplicados de nombres
        if range_name in [dn for dn in wb.defined_names]:
            range_name = range_name + f"_{cliente.cliente_id}"
        wb.defined_names[range_name] = DefinedName(range_name, attr_text=attr)
        cliente_range_map[cliente.cliente_id] = range_name

    # Guardar mapa cliente→range_name en hoja oculta para usarlo al importar
    ws_map = wb.create_sheet("_mapa_clientes")
    ws_map.sheet_state = "hidden"
    ws_map.append(["cliente_nombre", "range_name", "cliente_id"])
    for cliente in clientes:
        rn = cliente_range_map.get(cliente.cliente_id, "TODAS_RUTAS")
        ws_map.append([cliente.nombre, rn, cliente.cliente_id])

    # ── Hoja Importar ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Importar")
    wb.active = ws

    NUM_COLS = len(COLUMNAS)   # 14
    LAST_COL = get_column_letter(NUM_COLS)

    # ── Fila 1: Título ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{LAST_COL}1")
    ws["A1"] = (
        "TransitPro — Plantilla de importación  |  "
        "Columnas AZULES = IDA   |   Columnas VERDES = RETORNO (opcionales)   |   "
        "Usa los desplegables ▼ — no se necesitan IDs"
    )
    ws["A1"].font      = Font(bold=True, size=10)
    ws["A1"].fill      = PatternFill("solid", start_color=COLOR_INFO)
    ws["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # ── Fila 2: Instrucción ───────────────────────────────────────────────────
    ws.merge_cells(f"A2:{LAST_COL}2")
    ws["A2"] = (
        "⚠  Primero selecciona el Cliente (col. B) → la Ruta se filtra automáticamente. "
        "Para el RETORNO: indica si la misma unidad regresa (col. K). "
        "Si es otra unidad, selecciona Placa y Conductor de retorno (cols. L y M) y la Hora de Salida desde Planta (col. N)."
    )
    ws["A2"].font      = Font(italic=True, size=9, color="7F4F00")
    ws["A2"].fill      = PatternFill("solid", start_color="FFE699")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[2].height = 32

    # ── Fila 3: Separador de secciones IDA / RETORNO ──────────────────────────
    # Sección IDA (cols A-J)
    ws.merge_cells("A3:J3")
    ws["A3"] = "🚌  TRAMO DE IDA"
    ws["A3"].font      = Font(bold=True, size=10, color=COLOR_WHITE)
    ws["A3"].fill      = PatternFill("solid", start_color="1A5276")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Sección RETORNO (cols K-N)
    ws.merge_cells("K3:N3")
    ws["K3"] = "↩  RETORNO"
    ws["K3"].font      = Font(bold=True, size=10, color=COLOR_WHITE)
    ws["K3"].fill      = PatternFill("solid", start_color="0E6655")
    ws["K3"].alignment = Alignment(horizontal="center", vertical="center")

    # ── Fila 4: Encabezados de columna ────────────────────────────────────────
    COLOR_HDR_IDA     = "2D6A9F"   # Azul IDA
    COLOR_HDR_RET     = "17735A"   # Verde RETORNO
    COLOR_DATA_IDA_R  = "D6E8F7"   # Azul claro obligatorio
    COLOR_DATA_IDA_O  = "EBF1DE"   # Verde claro opcional IDA
    COLOR_DATA_RET    = "D1F2EB"   # Verde menta — retorno opcional

    for ci, (label, req) in enumerate(COLUMNAS, start=1):
        hdr_color = COLOR_HDR_IDA if ci <= 10 else COLOR_HDR_RET
        c = ws.cell(row=4, column=ci, value=label)
        c.font      = Font(bold=True, color=COLOR_WHITE, size=9)
        c.fill      = PatternFill("solid", start_color=hdr_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN
    ws.row_dimensions[4].height = 36

    # ── Filas de datos (5 en adelante) ────────────────────────────────────────
    ROW_DATA_START = 5
    ultima = ROW_DATA_START + NUM_FILAS - 1

    for row in range(ROW_DATA_START, ultima + 1):
        for ci, (_, req) in enumerate(COLUMNAS, start=1):
            c = ws.cell(row=row, column=ci)
            if ci <= 10:
                c.fill = PatternFill("solid", start_color=COLOR_DATA_IDA_R if req else COLOR_DATA_IDA_O)
            else:
                c.fill = PatternFill("solid", start_color=COLOR_DATA_RET)
            c.border = THIN
            if ci == 1:       c.number_format = "DD/MM/YYYY"
            if ci in (7, 8, 14): c.number_format = "HH:MM"

    # ── Anchos de columna ─────────────────────────────────────────────────────
    anchos = [14, 22, 40, 16, 30, 30, 13, 14, 28, 26,   # IDA
              14, 16, 30, 14]                             # RETORNO
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ── Dropdowns IDA ─────────────────────────────────────────────────────────
    n_veh  = max(len(vehiculos), 1)
    n_cond = max(len(conductores), 1)

    # Cliente (B)
    dv_cli = DataValidation(type="list", formula1="LISTA_CLIENTES",
                            allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cli)
    dv_cli.add(f"B{ROW_DATA_START}:B{ultima}")

    # Ruta (C) — INDIRECT por cliente
    formula_ruta = (
        '=INDIRECT(IF(B5="","TODAS_RUTAS",'
        'IFERROR(VLOOKUP(B5,_mapa_clientes!$A:$B,2,0),"TODAS_RUTAS")))'
    )
    for row in range(ROW_DATA_START, ultima + 1):
        f = formula_ruta.replace(f"B{ROW_DATA_START}", f"B{row}")
        dv_r = DataValidation(type="list", formula1=f, allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv_r)
        dv_r.add(f"C{row}")

    # Vehículo IDA (D)
    dv_veh = DataValidation(type="list", formula1=f"Vehiculos!$A$2:$A${n_veh+1}",
                            allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_veh)
    dv_veh.add(f"D{ROW_DATA_START}:D{ultima}")

    # Conductores IDA (E, F)
    dv_cond = DataValidation(type="list", formula1=f"Conductores!$A$2:$A${n_cond+1}",
                             allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cond)
    dv_cond.add(f"E{ROW_DATA_START}:F{ultima}")

    # Paradero (I)
    if paraderos:
        n_par = len(paraderos)
        dv_par = DataValidation(type="list", formula1=f"Paraderos!$A$2:$A${n_par+1}",
                                allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv_par)
        dv_par.add(f"I{ROW_DATA_START}:I{ultima}")

    # ── Dropdowns RETORNO ─────────────────────────────────────────────────────
    # Misma unidad (K): Sí / No
    dv_misma = DataValidation(type="list", formula1='"Sí,No"',
                              allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_misma)
    dv_misma.add(f"K{ROW_DATA_START}:K{ultima}")

    # Vehículo RETORNO (L) — mismo dropdown que IDA
    dv_veh_ret = DataValidation(type="list", formula1=f"Vehiculos!$A$2:$A${n_veh+1}",
                                allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_veh_ret)
    dv_veh_ret.add(f"L{ROW_DATA_START}:L{ultima}")

    # Conductor RETORNO (M)
    dv_cond_ret = DataValidation(type="list", formula1=f"Conductores!$A$2:$A${n_cond+1}",
                                 allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cond_ret)
    dv_cond_ret.add(f"M{ROW_DATA_START}:M{ultima}")

    # Fila de ejemplo (en fila 5, primera fila de datos)
    if rutas and vehiculos and conductores:
        ejm = [
            "14/04/2026",
            clientes[0].nombre if clientes else "",
            rutas[0].nombre,
            vehiculos[0].placa,
            conductores[0]["nombre_completo"],
            "",
            "06:00", "", "",
            "Ejemplo — borrar",
            # Retorno de ejemplo
            "Sí",      # K - misma unidad
            "",        # L - placa retorno (vacío porque es misma unidad)
            "",        # M - conductor retorno (vacío)
            "14:30",   # N - hora salida planta
        ]
        for ci, val in enumerate(ejm, start=1):
            c = ws.cell(row=ROW_DATA_START, column=ci, value=val)
            c.fill = PatternFill("solid", start_color="D9D9D9")
            c.font = Font(italic=True, color="595959", size=9)
            c.border = THIN

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Hojas de referencia ───────────────────────────────────────────────────────
def _hdr_row(ws, cols):
    ws.append(cols)
    for ci in range(1, len(cols) + 1):
        c = ws.cell(row=1, column=ci)
        c.font = Font(bold=True, color=COLOR_WHITE)
        c.fill = PatternFill("solid", start_color=COLOR_REF_HDR)
        c.alignment = Alignment(horizontal="center")

def _ref_clientes(ws, clientes):
    _hdr_row(ws, ["Nombre Cliente"])
    for c in clientes:
        ws.append([c.nombre])
    ws.column_dimensions["A"].width = 30

def _ref_vehiculos(ws, vehiculos):
    _hdr_row(ws, ["Placa", "Marca", "Modelo"])
    for v in vehiculos:
        ws.append([v.placa, v.marca or "", v.modelo or ""])
    for i, w in enumerate([14, 16, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _ref_conductores(ws, conductores):
    _hdr_row(ws, ["Nombre Completo", "Licencia"])
    for c in conductores:
        ws.append([c.get("nombre_completo", ""), c.get("emp_licencia", "") or ""])
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14

def _ref_paraderos(ws, paraderos):
    _hdr_row(ws, ["Nombre Paradero"])
    for p in paraderos:
        ws.append([p.nombre])
    ws.column_dimensions["A"].width = 35


# ─────────────────────────────────────────────────────────────────────────────
# IMPORTAR DESDE EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def importar_desde_excel(archivo: BinaryIO, sede_id: int, usuario_id: int, db: Session) -> dict:
    try:
        wb = load_workbook(archivo, data_only=True)
    except Exception as exc:
        return {"ok": False, "error": f"No se pudo leer el archivo: {exc}"}

    hoja = wb.get("Importar") or wb.get("importar")
    if hoja is None:
        return {"ok": False, "error": "El archivo no contiene la hoja 'Importar'."}

    # ── Catálogos para match ──────────────────────────────────────────────────
    rutas       = db.query(Ruta).filter(Ruta.sede_id == sede_id, Ruta.activa == True).all()
    vehiculos   = db.query(Vehiculo).filter(Vehiculo.sede_id == sede_id, Vehiculo.operativo == True).all()
    conductores = get_conductores_by_sede(sede_id, db)
    paraderos   = db.query(Paradero).filter(Paradero.sede_id == sede_id, Paradero.activo == True).all()
    clientes    = db.query(Cliente).filter(Cliente.activo == True).all()

    idx_rutas  = {r.nombre.strip().upper(): r for r in rutas}
    idx_placas = {v.placa.strip().upper(): v for v in vehiculos}
    idx_conds  = {c["nombre_completo"].strip().upper(): c for c in conductores}
    idx_parad  = {p.nombre.strip().upper(): p for p in paraderos}
    idx_cli    = {c.nombre.strip().upper(): c for c in clientes}

    resultados = []
    creados = errores = 0

    # Detectar fila de inicio de datos (puede ser fila 4 o 5 según versión plantilla)
    fila_inicio = 4
    for test_row in (4, 5):
        first_vals = list(hoja.iter_rows(min_row=test_row, max_row=test_row, values_only=True))
        if first_vals and any(v is not None for v in first_vals[0]):
            break
        fila_inicio = test_row + 1

    for fila_num, fila in enumerate(hoja.iter_rows(min_row=fila_inicio, values_only=True), start=fila_inicio):
        if all(v is None or str(v).strip() == "" for v in fila):
            continue

        vals = [(str(v).strip() if v is not None else "") for v in fila]
        # Saltar fila de ejemplo
        if any("ejemplo" in (v.lower()) for v in vals if v):
            continue

        # 14 columnas: 10 IDA + 4 RETORNO
        padded = (vals + [""] * 14)[:14]
        fecha_r, cli_r, ruta_r, placa_r, c1_r, c2_r, hi_r, hf_r, par_r, obs_r, \
        misma_r, placa_ret_r, cond_ret_r, hora_planta_r = padded

        # ── Fecha ──────────────────────────────────────────────────────────────
        fecha_val = _parsear_fecha(fecha_r)
        if not fecha_val:
            resultados.append({"fila": fila_num, "estado": "ERROR",
                "detalle": f"Fecha '{fecha_r}' no reconocida. Usa DD/MM/AAAA."})
            errores += 1; continue

        # ── Ruta (match exacto primero, luego parcial) ─────────────────────────
        if not ruta_r:
            resultados.append({"fila": fila_num, "estado": "ERROR", "detalle": "Ruta vacía."})
            errores += 1; continue
        ruta = idx_rutas.get(ruta_r.upper()) or \
               next((r for r in rutas if ruta_r.upper() in r.nombre.upper()), None)
        if not ruta:
            resultados.append({"fila": fila_num, "estado": "ERROR",
                "detalle": f"Ruta '{ruta_r}' no encontrada. Verifica el desplegable."})
            errores += 1; continue

        # Verificar que la ruta pertenece al cliente indicado (si se escribió uno)
        if cli_r:
            cliente_obj = idx_cli.get(cli_r.upper())
            if cliente_obj and ruta.cliente_id != cliente_obj.cliente_id:
                resultados.append({"fila": fila_num, "estado": "ERROR",
                    "detalle": f"La ruta '{ruta_r}' no pertenece al cliente '{cli_r}'."})
                errores += 1; continue

        # ── Vehículo ───────────────────────────────────────────────────────────
        if not placa_r:
            resultados.append({"fila": fila_num, "estado": "ERROR", "detalle": "Placa vacía."})
            errores += 1; continue
        vehiculo = idx_placas.get(placa_r.upper())
        if not vehiculo:
            resultados.append({"fila": fila_num, "estado": "ERROR",
                "detalle": f"Placa '{placa_r}' no encontrada o vehículo inoperativo."})
            errores += 1; continue

        # ── Conductor 1 ────────────────────────────────────────────────────────
        if not c1_r:
            resultados.append({"fila": fila_num, "estado": "ERROR", "detalle": "Conductor 1 vacío."})
            errores += 1; continue
        cond1 = idx_conds.get(c1_r.upper()) or \
                next((c for c in conductores if c1_r.upper() in c["nombre_completo"].upper()), None)
        if not cond1:
            resultados.append({"fila": fila_num, "estado": "ERROR",
                "detalle": f"Conductor '{c1_r}' no encontrado."})
            errores += 1; continue

        # ── Conductor 2 ────────────────────────────────────────────────────────
        cond2_id = None
        if c2_r:
            cond2 = idx_conds.get(c2_r.upper()) or \
                    next((c for c in conductores if c2_r.upper() in c["nombre_completo"].upper()), None)
            if not cond2:
                resultados.append({"fila": fila_num, "estado": "ERROR",
                    "detalle": f"Conductor 2 '{c2_r}' no encontrado."})
                errores += 1; continue
            cond2_id = cond2["emp_id"]

        # ── Horas ──────────────────────────────────────────────────────────────
        hora_ini = _parsear_hora(hi_r)
        if not hora_ini:
            resultados.append({"fila": fila_num, "estado": "ERROR",
                "detalle": f"Hora inicio '{hi_r}' inválida."})
            errores += 1; continue
        hora_fin = _parsear_hora(hf_r) if hf_r else None

        # ── Paradero ───────────────────────────────────────────────────────────
        paradero_id = None
        if par_r:
            p = idx_parad.get(par_r.upper()) or \
                next((x for x in paraderos if par_r.upper() in x.nombre.upper()), None)
            if p:
                paradero_id = p.paradero_id

        # ── Retorno ────────────────────────────────────────────────────────────
        misma_unidad     = misma_r.strip().lower() not in ("no", "n", "false", "0")
        ret_vehiculo_id  = None
        ret_conductor_id = None
        hora_salida_planta = _parsear_hora(hora_planta_r) if hora_planta_r else None

        if not misma_unidad:
            # Vehículo de retorno
            if placa_ret_r:
                veh_ret = idx_placas.get(placa_ret_r.upper())
                if not veh_ret:
                    resultados.append({"fila": fila_num, "estado": "ERROR",
                        "detalle": f"Placa retorno '{placa_ret_r}' no encontrada."})
                    errores += 1; continue
                ret_vehiculo_id = veh_ret.vehiculo_id
            # Conductor de retorno
            if cond_ret_r:
                cond_ret = idx_conds.get(cond_ret_r.upper()) or \
                           next((c for c in conductores if cond_ret_r.upper() in c["nombre_completo"].upper()), None)
                if not cond_ret:
                    resultados.append({"fila": fila_num, "estado": "ERROR",
                        "detalle": f"Conductor retorno '{cond_ret_r}' no encontrado."})
                    errores += 1; continue
                ret_conductor_id = cond_ret["emp_id"]

        # ── Crear y validar ────────────────────────────────────────────────────
        from sqlalchemy.orm import joinedload
        from ..services.programacion import validar_y_enriquecer
        from fastapi import HTTPException

        ruta_full = db.query(Ruta).options(joinedload(Ruta.sede)).filter(
            Ruta.ruta_id == ruta.ruta_id).first()
        try:
            schema_data = ServicioCreate(
                fecha=fecha_val, sede_id=sede_id, ruta_id=ruta.ruta_id,
                vehiculo_id=vehiculo.vehiculo_id, conductor_id=cond1["emp_id"],
                conductor2_id=cond2_id, hora_inicio=hora_ini,
                hora_fin_manual=hora_fin, paradero_origen_id=paradero_id,
                observaciones=obs_r or None,
                retorno_misma_unidad=misma_unidad,
                retorno_vehiculo_id=ret_vehiculo_id,
                retorno_conductor_id=ret_conductor_id,
                hora_salida_planta=hora_salida_planta,
            )
            extras = validar_y_enriquecer(schema_data, ruta_full, db)
        except HTTPException as e:
            resultados.append({"fila": fila_num, "estado": "ERROR", "detalle": e.detail})
            errores += 1; continue
        except Exception as e:
            resultados.append({"fila": fila_num, "estado": "ERROR", "detalle": str(e)})
            errores += 1; continue

        try:
            from ..models.servicio import Servicio
            sd = schema_data.model_dump()
            sd.update(extras)
            sd["creado_por"] = usuario_id
            sd["estado"]     = "PROGRAMADO"
            s = Servicio(**sd)
            db.add(s); db.flush()
            resultados.append({"fila": fila_num, "estado": "OK", "servicio_id": s.servicio_id,
                "detalle": f"{ruta.nombre} | {vehiculo.placa} | {hora_ini.strftime('%H:%M')}"})
            creados += 1
        except Exception as e:
            db.rollback()
            resultados.append({"fila": fila_num, "estado": "ERROR", "detalle": f"Error al guardar: {e}"})
            errores += 1

    if creados > 0:
        db.commit()

    return {"total_filas": creados + errores, "creados": creados,
            "errores": errores, "detalle": resultados}


# ── Parseo de fechas y horas ──────────────────────────────────────────────────
def _parsear_fecha(raw: str) -> date | None:
    if not raw: return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try: return datetime.strptime(raw.strip(), fmt).date()
        except ValueError: pass
    try:
        from datetime import timedelta
        return (datetime(1899, 12, 30) + timedelta(days=int(float(raw)))).date()
    except: pass
    return None


def _parsear_hora(raw: str) -> time | None:
    if not raw: return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try: return datetime.strptime(raw.strip(), fmt).time()
        except ValueError: pass
    try:
        f = float(raw)
        total = round(f * 86400)
        h, r = divmod(total, 3600); m, s = divmod(r, 60)
        return time(h % 24, m, s)
    except: pass
    return None
